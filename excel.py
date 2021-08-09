from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import load_workbook

wb = load_workbook("almacen.xlsx")
ws = wb.active
columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V",
           "W", "X", "Y", "Z"]


def ennumeracion():
    cuenta = int(input("Hasta que numero deseas enumerar? "))
    for x in range(cuenta):
        s = str((x + 1))
        n = columns[1] + s
        ws[n] = x + 1
ennumeracion()

def introduccion():
    entrada = str(input("entrada: "))
    locacion = str(input("localidad: "))
    ws[locacion] = entrada
    pregunta = str(input("quisieras continuar agregando? "))

    if (pregunta == 'si'):
        introduccion()
        wb.save("almacen.xlsx")
    else:
        wb.save("almacen.xlsx")
        quit()


def aricmetica():
    primer = ws["C1"].value
    segundo = ws["C2"].value
    primeros = int(primer)
    segundos = int(segundo)
    suma = primeros + segundos
    sumas = str(suma)
    ws["C3"] = sumas
    wb.save("almacen.xlsx")


def buscador(columns):
    n = 0
    buscar = str(input("palabra a buscar: "))

    for colubnas in columns:
        s = columns[n]
        for fila in range(100):
            decender_en_colugna = str(ws[(s + (str(fila + 1)))].value)
            if (decender_en_colugna == buscar):
                print("lo encontre, esta en: ", ws[(s + (str(fila + 1)))])

        n += 1


def borrador(columns):
    for m in range(100):
        s = str(m + 1)
        ws[columns[0] + s] = ""
        ws[columns[1] + s] = ""
        ws[columns[2] + s] = ""
    wb.save("almacen.xlsx")


def cambio_de_moneda():
    # este programa te permite ingresar un numero representando una cantidad de dinero en dolares y luego ir a una pagina de internet que hace conversion, cambiarlo a pesos dominicanos
    # y luego entrarlo en una celda de excel el numero ingresado y el resultado.
    hora_actual = 0
    minuto_actual = str(time.localtime().tm_min)
    hora = int(time.localtime().tm_hour)
    minutos = int(time.localtime().tm_min)
    if (hora > 12):
        hora_actual = hora - 12
    elif (minutos < 10):
        minuto_actual = "0" + str(minutos)
    PATH = r"C:\chromedriver.exe"
    cantidad_de_dinero_a_cambiar = int(input("Que cantidad de dinero desea cambiar? "))
    driver = webdriver.Chrome(PATH)
    driver.get("https://www.currency-calc.com/USD_DOP")
    driver.implicitly_wait(5)
    cantidad = driver.find_element_by_class_name("currency-field-input")
    cantidad.click()
    time.sleep(5)
    cantidad.send_keys(cantidad_de_dinero_a_cambiar)
    time.sleep(5)
    cantidad.send_keys(Keys.RETURN)
    resultado = driver.find_element_by_class_name("currency-field-result-number").text
    driver.close()
    ws[columns[1] + "1"] = "Dolares"
    ws[columns[2] + "1"] = "pesos"
    ws[columns[1] + "2"] = cantidad_de_dinero_a_cambiar
    ws[columns[2] + "2"] = resultado
    ws[columns[3] + "1"] = "Dia"
    ws[columns[4] + "1"] = "Mes"
    ws[columns[5] + "1"] = "Año"
    ws[columns[3] + "2"] = str(time.localtime().tm_mday)
    ws[columns[4] + "2"] = str(time.localtime().tm_mon)
    ws[columns[5] + "2"] = str(time.localtime().tm_year)
    ws[columns[6] + "1"] = "Hora registrada"
    ws[columns[6] + "2"] = str(hora_actual) + ":" + minuto_actual
    print("El resultado es ", resultado, " pesos dominicanos.")

    time.sleep(5)
    wb.save("almacen.xlsx")


wb.save("almacen.xlsx")
