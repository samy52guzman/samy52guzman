from selenium import webdriver
import time
from openpyxl import load_workbook
#Program that goes to the link given, take the numbers of the last powerball, and saves it in a excel file.
numeros = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V",
           "W", "X", "Y", "Z"]


def chrome():
    lista = []
    PATH = r"C:\chromedriver.exe"

    driver = webdriver.Chrome(PATH)
    driver.get("https://www.lotteryusa.com/powerball/")
    driver.implicitly_wait(2)
    primer = driver.find_element_by_class_name("c-result-card__result-list").text

    lista.append(primer)

    time.sleep(1)
    n = 0
    numeros_separados = []
    while n < (len(lista[0]) - 1):
        x = lista[0][n]
        if x == "\n" or x not in numeros:
            numeros_separados.append("")
        else:
            if lista[0][n + 1] != "\n":
                numeros_separados.append(x + lista[0][n + 1])
            elif lista[0][n + 1] == "\n" and lista[0][n - 1] not in numeros:
                numeros_separados.append(x)

        n += 1

    driver.close()
    excel(numeros_separados)
    return numeros_separados


def excel(numeros_separados):
    i = 1
    wb = load_workbook("PowerBall.xlsx")
    ws = wb.active
    ws["B1"] = "Numeros del Power ball"
    for m in numeros_separados:
        if m != "":
            ws[columns[i] + "2"] = m
            i += 1
    wb.save("PowerBall.xlsx")


chrome()
